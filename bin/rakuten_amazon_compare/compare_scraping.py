import datetime
import re
import time
import tkinter
import openpyxl

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import chromedriver_binary

date = datetime.datetime.now().strftime('%Y%m%d')
EXTENSION_PATH = 'quick_shop1.9.4.0.crx'
BASE_URL = 'https://www.amazon.co.jp/'
# usernameの部分をパソコン環境ごとに変更する
USER_CHROME_PATH = '--user-data-dir=/Users/username/Library/Application\ Support/Google/Chrome/Default/'


def main():
    search_text = search_form.get()

    workbook = openpyxl.load_workbook('rakuten_amazon.xlsx')
    i = 1

    options = Options()
    options.add_argument(USER_CHROME_PATH)
    options.add_extension(EXTENSION_PATH)
    driver = webdriver.Chrome(options=options)
    driver.get(BASE_URL)
    time.sleep(2)
    search_input = driver.find_element(By.ID, 'twotabsearchtextbox')
    search_input.send_keys(search_text)
    search_input.send_keys(Keys.ENTER)
    time.sleep(15)

    try:
        while True:
            soup = BeautifulSoup(driver.page_source.encode('utf-8'), 'html.parser')
            for item in soup.find_all('div', class_='s-widget-spacing-small'):
                amazon_title = item.find('h2').text.strip()
                amazon_price_container = item.find('span', class_='a-price-whole')
                if amazon_price_container:
                    amazon_price = int(re.sub('[^0-9]', '', amazon_price_container.text))
                else:
                    amazon_price = 'NO DATA'
                asin_container = item.find('div', class_='qs-asin')
                if asin_container:
                    asin = asin_container.contents[3].text
                else:
                    continue
                amazon_item_url = 'https://www.amazon.co.jp/dp/{}'.format(asin)
                driver.get(amazon_item_url)
                time.sleep(5)
                amazon_soup = BeautifulSoup(driver.page_source.encode('utf-8'), 'html.parser')

                jan_code = None
                jan_code_container = amazon_soup.find('div', class_='qs-jan')
                if jan_code_container:
                    if jan_code_container.contents[3].text.isdecimal():
                        jan_code = int(jan_code_container.contents[3].text)

                amazon_point_container = amazon_soup.find('div', id='points_feature_div')
                if amazon_point_container:
                    amazon_point_text = amazon_point_container.text
                    if amazon_point_text.isdecimal():
                        amazon_point = int(re.sub('[^0-9]', '', amazon_point_text[:amazon_point_text.find('p')]))
                    else:
                        amazon_point = 0
                else:
                    amazon_point = 0
                model_text = amazon_soup.find('div', class_='qs-model').contents[3].text
                if model_text != '':
                    model_code = model_text
                else:
                    model_code = 'NO DATA'
                amazon_item_models = amazon_soup.find_all('td', class_='a-span7 a-size-base')
                if len(amazon_item_models) >= 2:
                    amazon_item_model_text = amazon_item_models[1].text
                    amazon_item_model = amazon_item_model_text.replace(' ', '')
                else:
                    amazon_item_model = 'NO DATA'

                if jan_code:
                    rakuten_item_list_url = 'https://search.rakuten.co.jp/search/mall/{}/'.format(jan_code)
                    driver.get(rakuten_item_list_url)
                    time.sleep(2)
                    rakuten_list_soup = BeautifulSoup(driver.page_source.encode('utf-8'), 'html.parser')
                    rakuten_item = rakuten_list_soup.find('div', class_='searchresultitem')
                    if not rakuten_item:
                        rakuten_item_list_url = 'https://search.rakuten.co.jp/search/mall/{}/'.format(asin)
                        driver.get(rakuten_item_list_url)
                        time.sleep(2)
                        rakuten_list_soup = BeautifulSoup(driver.page_source.encode('utf-8'), 'html.parser')
                        rakuten_item = rakuten_list_soup.find('div', class_='searchresultitem')
                else:
                    jan_code = 'NO DATA'
                    rakuten_item_list_url = 'https://search.rakuten.co.jp/search/mall/{}/'.format(asin)
                    driver.get(rakuten_item_list_url)
                    time.sleep(2)
                    rakuten_list_soup = BeautifulSoup(driver.page_source.encode('utf-8'), 'html.parser')
                    rakuten_item = rakuten_list_soup.find('div', class_='searchresultitem')

                if rakuten_item:
                    rakuten_title = rakuten_item.find('h2').text.strip()
                    rakuten_price = int(re.sub('[^0-9]', '', rakuten_item.find('span', class_='important').text))
                    rakuten_item_url = rakuten_item.find('a').get('href')

                else:
                    rakuten_title = 'NO DATA'
                    rakuten_price = 'NO DATA'
                    rakuten_item_url = 'NO DATA'

                sheet = workbook['シート1']
                sheet.cell(column=1, row=i + 1, value=asin)
                sheet.cell(column=2, row=i + 1, value=jan_code)
                sheet.cell(column=3, row=i + 1, value=model_code)
                sheet.cell(column=4, row=i + 1, value=amazon_item_model)
                sheet.cell(column=5, row=i + 1, value=amazon_item_url)
                sheet.cell(column=6, row=i + 1, value=amazon_title)
                sheet.cell(column=7, row=i + 1, value=amazon_price)
                sheet.cell(column=8, row=i + 1, value=amazon_point)
                sheet.cell(column=9, row=i + 1, value=rakuten_item_url)
                sheet.cell(column=10, row=i + 1, value=rakuten_title)
                sheet.cell(column=11, row=i + 1, value=rakuten_price)

                i = i + 1

            if hasattr(soup.find('a', class_='s-pagination-next'), 'text'):
                next_url = soup.find('a', class_='s-pagination-next').get('href')
                driver.get('https://www.amazon.co.jp{}'.format(next_url))
                time.sleep(15)
            else:
                i = i - 1
                print('正常に終了しました')
                print('{}件の商品を取得しました'.format(i))
                break

    except Exception as e:
        print('エラーが発生しました')
        print('{}件の商品を取得しました'.format(i))

    file_name = '{}_rakuten_amazon_compare.xlsx'.format(date)
    workbook.save('date/{}'.format(file_name))
    workbook.close()


if __name__ == '__main__':
    window = tkinter.Tk()
    window.title('楽天×Amazon')
    window.geometry("650x150")

    search_label = tkinter.Label(text='検索ワード')
    search_label.place(x=50, y=50)

    search_form = tkinter.Entry(width=30)
    search_form.place(x=130, y=50)

    btn = tkinter.Button(window, text='実行', command=main)
    btn.place(x=450, y=50, width=100, height=30)

    window.mainloop()
